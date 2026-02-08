"""
Passport OCR API - FastAPI Backend
Extracts passport data from scanned images using OCR.
"""

import io
import json
import base64
from typing import List

from fastapi import FastAPI, UploadFile, File, HTTPException, Depends, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from PIL import Image
import openpyxl
from openpyxl.styles import Alignment, Border, Side

from .config import CORS_ORIGINS
from .auth import verify_password, create_access_token, verify_token
from .ocr_service import process_passport_image, rotate_image_arbitrary
import numpy as np
import cv2

app = FastAPI(
    title="Passport OCR API",
    description="Extract passport data from scanned images",
    version="1.0.0"
)

# CORS configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],
)


# Request/Response Models
class LoginRequest(BaseModel):
    username: str
    password: str


class LoginResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"


class PassportData(BaseModel):
    first_name: str = ""
    middle_name: str = ""
    last_name: str = ""
    gender: str = ""
    date_of_birth: str = ""
    nationality: str = ""
    passport_number: str = ""
    checkout_date: str = ""  # Check-out date from UI
    phone_number: str = ""  # Phone number from UI
    thumbnail: str = ""  # Base64 encoded thumbnail
    full_image: str = ""  # Base64 encoded full/preview image
    confidence: float = 0.0  # OCR confidence score
    low_confidence_fields: List[str] = []  # Fields that need manual review


class OCRResponse(BaseModel):
    success: bool
    passports: List[PassportData]
    message: str = ""


# Routes
@app.get("/")
async def root():
    """Health check endpoint."""
    return {"status": "ok", "service": "Passport OCR API"}


@app.get("/health")
async def health():
    """Health check for deployment."""
    return {"status": "healthy"}


@app.post("/api/login", response_model=LoginResponse)
async def login(request: LoginRequest):
    """Authenticate user and return JWT token."""
    if request.username != "admin" or not verify_password(request.password):
        raise HTTPException(
            status_code=401,
            detail="Invalid username or password"
        )

    access_token = create_access_token(data={"sub": request.username})
    return LoginResponse(access_token=access_token)


@app.post("/api/ocr", response_model=OCRResponse)
async def extract_passport_data(
    files: List[UploadFile] = File(...),
    rotation: float = Form(default=0),
    _: dict = Depends(verify_token)
):
    """
    Process uploaded passport images and extract data.
    Supports multiple files and multiple passports per image.
    Optional rotation parameter (degrees) for manual image correction.
    """
    all_passports = []

    for file in files:
        # Validate file type
        if not file.content_type or not any(
            t in file.content_type.lower()
            for t in ['image/jpeg', 'image/png', 'image/jpg', 'application/pdf']
        ):
            continue

        try:
            # Read file content
            content = await file.read()

            # Handle PDF conversion if needed
            if file.content_type == 'application/pdf':
                continue

            # Apply rotation if specified and generate images
            if rotation != 0:
                rotated_content = apply_rotation_to_image(content, rotation)
                thumbnail = create_thumbnail(rotated_content)
                full_image = create_preview_image(rotated_content)
                passports = process_passport_image(rotated_content, rotation_angle=0)
            else:
                thumbnail = create_thumbnail(content)
                full_image = create_preview_image(content)
                passports = process_passport_image(content)

            # Add images to each passport found
            for passport in passports:
                passport['thumbnail'] = thumbnail
                passport['full_image'] = full_image
                all_passports.append(PassportData(**passport))

            # If no passports found, add empty entry with images
            if not passports:
                all_passports.append(PassportData(
                    thumbnail=thumbnail,
                    full_image=full_image,
                    low_confidence_fields=["first_name", "last_name", "passport_number"]
                ))

        except Exception as e:
            print(f"Error processing file {file.filename}: {e}")
            continue

    if not all_passports:
        return OCRResponse(
            success=False,
            passports=[],
            message="No passport data could be extracted. Please ensure images are clear and contain valid passports."
        )

    return OCRResponse(
        success=True,
        passports=all_passports,
        message=f"Successfully processed {len(all_passports)} passport(s)"
    )


def apply_rotation_to_image(image_bytes: bytes, angle: float) -> bytes:
    """Apply rotation to image and return new bytes."""
    nparr = np.frombuffer(image_bytes, np.uint8)
    image = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

    if image is None:
        return image_bytes

    rotated = rotate_image_arbitrary(image, angle)

    # Encode back to bytes
    _, buffer = cv2.imencode('.jpg', rotated, [cv2.IMWRITE_JPEG_QUALITY, 95])
    return buffer.tobytes()


def create_thumbnail(image_bytes: bytes, max_size: int = 100) -> str:
    """Create a base64 encoded thumbnail from image bytes."""
    try:
        image = Image.open(io.BytesIO(image_bytes))
        image.thumbnail((max_size, max_size))

        # Convert to RGB if necessary
        if image.mode in ('RGBA', 'P'):
            image = image.convert('RGB')

        buffer = io.BytesIO()
        image.save(buffer, format='JPEG', quality=70)
        buffer.seek(0)

        return base64.b64encode(buffer.getvalue()).decode('utf-8')
    except Exception:
        return ""


def create_preview_image(image_bytes: bytes, max_size: int = 800) -> str:
    """Create a base64 encoded preview image (larger than thumbnail)."""
    try:
        image = Image.open(io.BytesIO(image_bytes))
        image.thumbnail((max_size, max_size))

        # Convert to RGB if necessary
        if image.mode in ('RGBA', 'P'):
            image = image.convert('RGB')

        buffer = io.BytesIO()
        image.save(buffer, format='JPEG', quality=85)
        buffer.seek(0)

        return base64.b64encode(buffer.getvalue()).decode('utf-8')
    except Exception:
        return ""


@app.post("/api/export")
async def export_to_excel(
    excel_file: UploadFile = File(...),
    passports_json: str = Form(...),
    _: dict = Depends(verify_token)
):
    """Append passport data to an existing Excel file."""
    # Parse passport data from JSON string
    try:
        passports_raw = json.loads(passports_json)
        passports = [PassportData(**p) for p in passports_raw]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid passport data: {e}")

    # Load the uploaded Excel file
    try:
        file_content = await excel_file.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Cannot read Excel file: {e}")

    # Define styles for new rows
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Find the actual next empty row by scanning column A (First Name)
    # ws.max_row can be inflated by empty formatted rows in templates
    next_row = 2  # Start after header row
    for row_idx in range(2, ws.max_row + 2):
        cell_value = ws.cell(row=row_idx, column=1).value
        if cell_value is None or str(cell_value).strip() == "":
            next_row = row_idx
            break
    # Append data rows
    for i, passport in enumerate(passports):
        row = next_row + i

        # Column order: First Name, Middle Name, Last Name, Gender, Passport No., Nationality, Birth Date, Checkout Date, Phone No.
        data = [
            passport.first_name,
            passport.middle_name,
            passport.last_name,
            passport.gender,
            passport.passport_number,
            passport.nationality,
            passport.date_of_birth,
            passport.checkout_date,
            passport.phone_number
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Return the modified file with original filename
    original_filename = excel_file.filename or "passport_data.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={original_filename}"}
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
